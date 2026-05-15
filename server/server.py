#!/usr/bin/env python3
"""운송 대출·매입 관리 서버"""

from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os, sys

app = Flask(__name__, static_folder='.')
CORS(app)

# ── 경로 설정: 스크립트 위치 기준 ──
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, '..', 'excel')
EXCEL_PATH = os.path.join(EXCEL_DIR, 'datalist.xlsx')

HEADERS = ['번호','등록일','거래처명','차량번호','운전자명','대출구분','매입구분',
           '대출금액','매입금액','이자율','만기일','비고','상태']
COL_WIDTHS = [6,12,15,12,12,10,10,14,14,8,12,20,8]

def ensure_excel():
    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = '대출매입관리'
        hf = PatternFill('solid', start_color='1F4E79')
        hfont = Font(bold=True, color='FFFFFF', name='맑은 고딕', size=10)
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for i,(h,w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = hfont; c.fill = hf; c.alignment = center; c.border = border
            ws.column_dimensions[chr(64+i)].width = w
        ws.row_dimensions[1].height = 25
        ws.freeze_panes = 'A2'
        wb.save(EXCEL_PATH)
        print(f'✅ Excel 파일 생성: {EXCEL_PATH}')

def row_border(ws, row_idx):
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    for col in range(1, len(HEADERS)+1):
        cell = ws.cell(row=row_idx, column=col)
        cell.border = border
        if col in [1,2,4,6,7,11,13]:
            cell.alignment = center
        if col in [8,9]:
            cell.number_format = '#,##0'
        if col == 10:
            cell.number_format = '0.00'

@app.route('/')
def index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/api/list')
def api_list():
    ensure_excel()
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        rec = {}
        for i, h in enumerate(HEADERS):
            v = row[i] if i < len(row) else None
            if hasattr(v, 'strftime'): v = v.strftime('%Y-%m-%d')
            rec[h] = '' if v is None else v
        records.append(rec)
    return jsonify(records)

@app.route('/api/save', methods=['POST'])
def api_save():
    ensure_excel()
    d = request.json
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    max_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            try: max_num = max(max_num, int(row[0]))
            except: pass
    new_num = max_num + 1
    today = datetime.now().strftime('%Y-%m-%d')
    row_data = [
        new_num, d.get('등록일', today), d.get('거래처명',''),
        d.get('차량번호',''), d.get('운전자명',''), d.get('대출구분',''), d.get('매입구분',''),
        float(d['대출금액']) if d.get('대출금액') else 0,
        float(d['매입금액']) if d.get('매입금액') else 0,
        float(d['이자율']) if d.get('이자율') else 0,
        d.get('만기일',''), d.get('비고',''), d.get('상태','정상'),
    ]
    nr = ws.max_row + 1
    for col, val in enumerate(row_data, 1):
        ws.cell(row=nr, column=col, value=val)
    row_border(ws, nr)
    wb.save(EXCEL_PATH)
    print(f'  저장: No.{new_num} {d.get("거래처명","")}')
    return jsonify({'success': True, 'num': new_num})

@app.route('/api/update', methods=['POST'])
def api_update():
    ensure_excel()
    d = request.json
    target = int(d.get('번호'))
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == target:
            row_data = [
                target, d.get('등록일',''), d.get('거래처명',''),
                d.get('차량번호',''), d.get('운전자명',''), d.get('대출구분',''), d.get('매입구분',''),
                float(d['대출금액']) if d.get('대출금액') else 0,
                float(d['매입금액']) if d.get('매입금액') else 0,
                float(d['이자율']) if d.get('이자율') else 0,
                d.get('만기일',''), d.get('비고',''), d.get('상태','정상'),
            ]
            for col, val in enumerate(row_data, 1):
                row[col-1].value = val
            row_border(ws, row[0].row)
            break
    wb.save(EXCEL_PATH)
    print(f'  수정: No.{target}')
    return jsonify({'success': True})

@app.route('/api/delete', methods=['POST'])
def api_delete():
    ensure_excel()
    target = int(request.json.get('번호'))
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == target:
            ws.delete_rows(row[0].row)
            break
    wb.save(EXCEL_PATH)
    print(f'  삭제: No.{target}')
    return jsonify({'success': True})

if __name__ == '__main__':
    ensure_excel()
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 5050
    print(f'\n🚛 운송 대출·매입 관리 서버 시작')
    print(f'   URL  : http://localhost:{port}')
    print(f'   Excel: {os.path.abspath(EXCEL_PATH)}\n')
    app.run(host='0.0.0.0', port=port, debug=False)
